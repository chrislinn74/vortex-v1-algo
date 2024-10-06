import pandas as pd
from openpyxl import load_workbook
import os
import pybaseball as pb

# Team abbreviation to full name mapping
TEAM_ABBR_TO_NAME = {
    'ARI': 'Arizona Diamondbacks', 'ATL': 'Atlanta Braves', 'BAL': 'Baltimore Orioles',
    'BOS': 'Boston Red Sox', 'CHC': 'Chicago Cubs', 'CHW': 'Chicago White Sox',
    'CIN': 'Cincinnati Reds', 'CLE': 'Cleveland Guardians', 'COL': 'Colorado Rockies',
    'DET': 'Detroit Tigers', 'HOU': 'Houston Astros', 'KCR': 'Kansas City Royals',
    'LAA': 'Los Angeles Angels', 'LAD': 'Los Angeles Dodgers', 'MIA': 'Miami Marlins',
    'MIL': 'Milwaukee Brewers', 'MIN': 'Minnesota Twins', 'NYM': 'New York Mets',
    'NYY': 'New York Yankees', 'OAK': 'Oakland Athletics', 'PHI': 'Philadelphia Phillies',
    'PIT': 'Pittsburgh Pirates', 'SDP': 'San Diego Padres', 'SFG': 'San Francisco Giants',
    'SEA': 'Seattle Mariners', 'STL': 'St. Louis Cardinals', 'TBR': 'Tampa Bay Rays',
    'TEX': 'Texas Rangers', 'TOR': 'Toronto Blue Jays', 'WSN': 'Washington Nationals'
}

# Lists of good and inverse metrics
GOOD_METRICS = ['run_differential', 'OPS', 'SLG', 'OBP', 'XBH', 'Hits', 'ISO', 'RAR', 'RBI', 'LOB%']
INVERSE_METRICS = ['ERA', 'FIP', 'WHIP', 'ERA-', 'FIP-', 'R', 'ER']

def save_to_excel(data, output_file):
    """Saves matchup data to an Excel sheet without overwriting existing formatting."""
    df = pd.DataFrame([data])

    if os.path.exists(output_file):
        # Load existing workbook
        book = load_workbook(output_file)
        
        # Get the active sheet
        sheet = book.active
        
        # Find the next empty row
        next_row = sheet.max_row + 1
        
        # Write new data
        for col, value in enumerate(df.iloc[0], start=1):
            sheet.cell(row=next_row, column=col, value=value)
        
        # Save the workbook
        book.save(output_file)
    else:
        # If file doesn't exist, create a new one
        df.to_excel(output_file, index=False)

def get_team_win_percentage(team, year):
    try:
        standings_list = pb.standings(year)
        standings = pd.concat(standings_list)
        team_name = TEAM_ABBR_TO_NAME.get(team, team)
        
        team_row = standings[standings['Tm'] == team_name]
        if not team_row.empty:
            wins = int(team_row['W'].values[0])
            losses = int(team_row['L'].values[0])
            games = wins + losses
            if games > 0:
                win_pct = wins / games
                return win_pct
            else:
                return 0.000
        else:
            print(f"Team '{team}' ({team_name}) not found in standings.")
            return None
    except Exception as e:
        print(f"Error fetching win percentage: {e}")
        return None